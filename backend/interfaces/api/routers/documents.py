"""Document Intelligence Pipeline — REST API.

Endpoints:
  POST   /documents/sources            → Neue Quelle registrieren
  GET    /documents/sources            → Alle Quellen auflisten
  PATCH  /documents/sources/{id}       → Quelle aktualisieren
  DELETE /documents/sources/{id}       → Quelle löschen
  POST   /documents/sources/{id}/ingest → Manuelle Ingestion triggern
  GET    /documents/files              → Alle verarbeiteten Dokumente
  GET    /documents/files/{id}         → Dokument-Detail mit AI-Extrakten
  POST   /documents/ingest-all         → Alle aktiven Quellen ingesten
"""

from __future__ import annotations

import asyncio
import os
import uuid
from datetime import UTC, datetime
from typing import Any

import structlog
from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from infrastructure.persistence.models.document_pipeline import (
    DocumentFileModel,
    DocumentSourceModel,
)
from application.rag.document_ingestion import (
    ingest_all_active_sources,
    ingest_source,
    ingest_uploaded_file_fast,
    process_document_background,
    reclassify_document,
    reclassify_all_documents,
    reanalyze_document,
    reparse_document,
    rechunk_document,
    reextract_metrics,
)
from application.rag.document_classifier import classify_document
from application.rag.embedder import embed_query
from infrastructure.persistence.models.rag_documents import RagDocumentModel
from infrastructure.persistence.models.company_intelligence import CompanyMetricModel, CompanySignalModel
from domain.user import User
from interfaces.api.deps import get_current_user, get_db
from shared.security import decode_token
from infrastructure.llm.deps import get_llm_provider
from application.ports.llm import Message

logger = structlog.get_logger(__name__)
router = APIRouter(prefix="/documents", tags=["Document Intelligence"])


# ── Schemas ───────────────────────────────────────────────────────────────────

class DocumentSourceCreate(BaseModel):
    supplier_id: str | None = None
    company_name: str | None = None
    doc_type: str  # annual_report | sustainability_report | audit_report | csrd_report | csddd_disclosure | sector_risk
    source_url: str
    schedule: str = "monthly"  # daily | weekly | monthly | manual


class DocumentSourceUpdate(BaseModel):
    company_name: str | None = None
    source_url: str | None = None
    schedule: str | None = None
    is_active: bool | None = None


class DocumentSourceOut(BaseModel):
    id: str
    organization_id: str
    supplier_id: str | None
    company_name: str | None
    doc_type: str
    source_url: str
    schedule: str
    is_active: bool
    last_fetched_at: datetime | None
    last_status: str | None
    last_error: str | None
    created_at: datetime
    updated_at: datetime

    model_config = {"from_attributes": True}


class DocumentFileOut(BaseModel):
    id: str
    organization_id: str
    source_id: str
    supplier_id: str | None
    doc_type: str
    title: str | None
    company_name: str | None
    report_year: int | None
    language: str | None
    file_url: str | None
    pages: int | None
    chunks_count: int | None
    esg_score: float | None
    summary: str | None
    extracted_risks: Any | None
    extracted_targets: Any | None
    extracted_commitments: Any | None
    extracted_kpis: Any | None
    status: str
    review_status: str
    copilot_hidden: bool
    classification_confidence: float | None
    error_msg: str | None
    created_at: datetime
    updated_at: datetime

    model_config = {"from_attributes": True}


class ReviewChunkOut(BaseModel):
    id: str
    content: str
    chunk_level: str
    doc_class: str | None
    page_number: int | None = None
    excluded_from_index: bool = False

class ReviewMetricOut(BaseModel):
    id: str
    metric_type: str
    value: float
    unit: str
    year: int
    period: str
    confidence: str
    confidence_pct: int | None = None
    page_number: int | None = None
    scope: str | None = None

class ReviewSignalOut(BaseModel):
    id: str
    signal_type: str
    dimension: str
    direction: str
    severity: str
    description: str
    year: int | None

class ReviewAuditEntry(BaseModel):
    id: str
    user_id: str
    action: str
    field: str | None
    old_value: str | None
    new_value: str | None
    created_at: datetime

class ReviewDataOut(BaseModel):
    id: str
    doc_type: str
    company_name: str | None
    report_year: int | None
    title: str | None
    language: str | None
    pages: int | None
    chunks_count: int | None
    esg_score: float | None
    summary: str | None
    status: str
    review_status: str
    review_notes: str | None
    parsed_text: str | None
    extracted_kpis: Any | None
    extracted_risks: Any | None
    extracted_targets: Any | None
    extracted_commitments: Any | None
    has_pdf: bool
    copilot_hidden: bool
    classification_confidence: float | None
    classification_alternatives: Any | None
    classification_evidence: list[str] | None
    created_at: datetime
    updated_at: datetime
    chunks: list[ReviewChunkOut]
    metrics: list[ReviewMetricOut]
    signals: list[ReviewSignalOut]
    audit_log: list[ReviewAuditEntry]

class ClassificationUpdate(BaseModel):
    doc_type: str | None = None
    company_name: str | None = None
    report_year: int | None = None

class KpiUpdate(BaseModel):
    kpis: dict[str, Any]

class ApproveRequest(BaseModel):
    notes: str | None = None

class ChunkContentUpdate(BaseModel):
    content: str

class TestRetrievalRequest(BaseModel):
    query: str
    top_k: int | None = None
    min_sim: float | None = None


# ── Sources ───────────────────────────────────────────────────────────────────

@router.post("/sources", response_model=DocumentSourceOut, status_code=status.HTTP_201_CREATED)
async def create_source(
    payload: DocumentSourceCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    org_id = user.organization_id
    # Upload-only sources must not be crawled by ingest_all
    is_active = not payload.source_url.startswith("upload://")
    source = DocumentSourceModel(
        id=str(uuid.uuid4()),
        organization_id=org_id,
        supplier_id=payload.supplier_id,
        company_name=payload.company_name,
        doc_type=payload.doc_type,
        source_url=payload.source_url,
        schedule=payload.schedule,
        is_active=is_active,
        created_at=datetime.now(UTC),
        updated_at=datetime.now(UTC),
    )
    db.add(source)
    await db.flush()
    logger.info("documents.source_created", org=org_id, source_id=source.id, doc_type=payload.doc_type)
    return source


@router.get("/sources", response_model=list[DocumentSourceOut])
async def list_sources(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    org_id = user.organization_id
    stmt = select(DocumentSourceModel).where(
        DocumentSourceModel.organization_id == org_id
    ).order_by(DocumentSourceModel.created_at.desc())
    result = await db.execute(stmt)
    return result.scalars().all()


@router.patch("/sources/{source_id}", response_model=DocumentSourceOut)
async def update_source(
    source_id: str,
    payload: DocumentSourceUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    org_id = user.organization_id
    source = await _get_source_or_404(source_id, org_id, db)
    if payload.company_name is not None:
        source.company_name = payload.company_name
    if payload.source_url is not None:
        source.source_url = payload.source_url
    if payload.schedule is not None:
        source.schedule = payload.schedule
    if payload.is_active is not None:
        source.is_active = payload.is_active
    source.updated_at = datetime.now(UTC)
    await db.flush()
    return source


@router.delete("/sources/{source_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_source(
    source_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    org_id = user.organization_id
    source = await _get_source_or_404(source_id, org_id, db)
    await db.delete(source)
    await db.flush()


@router.post("/sources/{source_id}/ingest")
async def trigger_ingest(
    source_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    org_id = user.organization_id
    source = await _get_source_or_404(source_id, org_id, db)
    stats = await ingest_source(source, db)
    return {"source_id": source_id, "stats": stats}


@router.post("/classify")
async def classify_document_endpoint(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
):
    """Auto-detect document type, year and title from uploaded file using heuristics + Groq."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="Empty file")
    if len(content) > 100 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 100 MB)")

    fn_lower = (file.filename or "").lower()
    content_type = "pdf" if fn_lower.endswith(".pdf") else "html"

    result = await classify_document(content, content_type, file.filename)
    return {
        "filename": file.filename,
        "doc_type": result.get("doc_type", "annual_report"),
        "report_year": result.get("report_year"),
        "title": result.get("title"),
        "confidence_source": result.get("confidence_source", "fallback"),
    }


@router.post("/sources/{source_id}/upload")
async def upload_document(
    source_id: str,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    report_year: int | None = None,
    title: str | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Upload a PDF or HTML file. Parses immediately, analysis + indexing run in background."""
    org_id = user.organization_id
    source = await _get_source_or_404(source_id, org_id, db)

    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="Empty file")
    if len(content) > 100 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 100 MB)")

    fn_lower = (file.filename or "").lower()
    content_type = "pdf" if fn_lower.endswith(".pdf") else "html"

    logger.info("documents.upload", org=org_id, source_id=source_id, filename=file.filename, size=len(content))

    stats = await ingest_uploaded_file_fast(
        source=source,
        content=content,
        content_type=content_type,
        filename=file.filename,
        session=db,
        report_year=report_year,
        title_override=title,
    )

    bg_args = stats.pop("background_args", None)
    if bg_args:
        # Commit NOW so the background task's new DB session can see the document.
        # FastAPI runs BackgroundTasks before yield-dependency cleanup (session.commit),
        # causing a race where the bg task opens a new session and finds nothing.
        await db.commit()
        background_tasks.add_task(process_document_background, **bg_args)

    return {"source_id": source_id, "filename": file.filename, "stats": stats}


# ── Files ─────────────────────────────────────────────────────────────────────

@router.get("/files", response_model=list[DocumentFileOut])
async def list_files(
    doc_type: str | None = None,
    supplier_id: str | None = None,
    status: str | None = None,
    limit: int = 200,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    org_id = user.organization_id
    stmt = select(DocumentFileModel).where(
        DocumentFileModel.organization_id == org_id
    )
    if doc_type:
        stmt = stmt.where(DocumentFileModel.doc_type == doc_type)
    if supplier_id:
        stmt = stmt.where(DocumentFileModel.supplier_id == supplier_id)
    if status:
        stmt = stmt.where(DocumentFileModel.status == status)
    stmt = stmt.order_by(DocumentFileModel.created_at.desc()).limit(min(limit, 500))
    result = await db.execute(stmt)
    return result.scalars().all()


# ── Requeue / Cleanup / Reclassify — must be BEFORE /files/{file_id} ────────

@router.post("/files/reclassify-all", status_code=status.HTTP_200_OK)
async def reclassify_all_files(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Re-run Groq classifier on all 'done' documents using stored chunk text."""
    org_id = user.organization_id
    result = await reclassify_all_documents(org_id, db)
    return result


@router.post("/files/{file_id}/reclassify", status_code=status.HTTP_200_OK)
async def reclassify_one_file(
    file_id: str,
    model: str | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Re-run Groq classifier. Pass model=groq:llama-3.1-8b-instant to override."""
    org_id = user.organization_id
    result = await reclassify_document(file_id, org_id, db, model=model)
    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])
    await db.execute(
        text("INSERT INTO document_review_log (id, doc_file_id, organization_id, user_id, action, field, old_value, new_value) VALUES (:id, :fid, :org, :uid, :action, :field, :old, :new)"),
        {"id": str(uuid.uuid4()), "fid": file_id, "org": org_id, "uid": user.id,
         "action": "reclassify", "field": "doc_type",
         "old": result.get("old_doc_type"), "new": result.get("new_doc_type")},
    )
    return result


class ReanalyzeRequest(BaseModel):
    extra_context: str | None = None

@router.post("/files/{file_id}/reanalyze", status_code=status.HTTP_200_OK)
async def reanalyze_one_file(
    file_id: str,
    model: str | None = None,
    payload: ReanalyzeRequest = ReanalyzeRequest(),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Re-run LLM analysis. Pass model to override, extra_context to add reviewer hint."""
    org_id = user.organization_id
    result = await reanalyze_document(file_id, org_id, db, model=model, extra_context=payload.extra_context)
    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])
    await db.execute(
        text("INSERT INTO document_review_log (id, doc_file_id, organization_id, user_id, action, field, old_value, new_value) VALUES (:id, :fid, :org, :uid, :action, :field, :old, :new)"),
        {"id": str(uuid.uuid4()), "fid": file_id, "org": org_id, "uid": user.id,
         "action": "reanalyze", "field": "summary",
         "old": None, "new": f"model={model or 'default'} kpis={result.get('kpi_count', 0)}" + (f" hint={payload.extra_context[:40]}" if payload.extra_context else "")},
    )
    return result


@router.post("/files/{file_id}/reparse", status_code=status.HTTP_200_OK)
async def reparse_one_file(
    file_id: str,
    parse_engine: str = "docling",
    ocr_enabled: bool = True,
    extract_tables: bool = True,
    describe_pictures: bool = False,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Re-parse PDF with selected engine and options."""
    result = await reparse_document(
        file_id, user.organization_id, db,
        parse_engine=parse_engine,
        ocr_enabled=ocr_enabled,
        extract_tables=extract_tables,
        describe_pictures=describe_pictures,
    )
    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])
    await db.execute(
        text("INSERT INTO document_review_log (id, doc_file_id, organization_id, user_id, action, field, old_value, new_value) VALUES (:id, :fid, :org, :uid, :action, :field, :old, :new)"),
        {"id": str(uuid.uuid4()), "fid": file_id, "org": user.organization_id, "uid": user.id,
         "action": "reparse", "field": "parsed_text",
         "old": None, "new": f"engine={parse_engine} ocr={ocr_enabled} tables={extract_tables} pictures={describe_pictures} pages={result.get('pages')} chars={result.get('chars')}"},
    )
    return result


@router.post("/files/{file_id}/rechunk", status_code=status.HTTP_200_OK)
async def rechunk_one_file(
    file_id: str,
    chunk_size: int = 800,
    chunk_overlap: int = 80,
    chunk_strategy: str = "sliding_window",
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Delete and re-create all chunks / embeddings for a document."""
    result = await rechunk_document(file_id, user.organization_id, db, chunk_size=chunk_size, chunk_overlap=chunk_overlap, chunk_strategy=chunk_strategy)
    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])
    await db.execute(
        text("INSERT INTO document_review_log (id, doc_file_id, organization_id, user_id, action, field, old_value, new_value) VALUES (:id, :fid, :org, :uid, :action, :field, :old, :new)"),
        {"id": str(uuid.uuid4()), "fid": file_id, "org": user.organization_id, "uid": user.id,
         "action": "rechunk", "field": "chunks",
         "old": None, "new": f"strategy={chunk_strategy} size={chunk_size} overlap={chunk_overlap} chunks={result.get('chunks_added')}"},
    )
    return result


@router.post("/files/{file_id}/reextract-metrics", status_code=status.HTTP_200_OK)
async def reextract_metrics_one_file(
    file_id: str,
    model: str | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Re-run metric/signal extraction. Pass model to override org default."""
    result = await reextract_metrics(file_id, user.organization_id, db, model=model)
    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])
    await db.execute(
        text("INSERT INTO document_review_log (id, doc_file_id, organization_id, user_id, action, field, old_value, new_value) VALUES (:id, :fid, :org, :uid, :action, :field, :old, :new)"),
        {"id": str(uuid.uuid4()), "fid": file_id, "org": user.organization_id, "uid": user.id,
         "action": "reextract_metrics", "field": "metrics",
         "old": None, "new": f"model={model or 'default'} metrics={result.get('metrics', 0)} signals={result.get('signals', 0)}"},
    )
    return result


@router.delete("/files/queued", status_code=status.HTTP_200_OK)
async def delete_queued_files(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Delete stuck 'queued' or 'error' files with no chunks so they can be re-uploaded."""
    org_id = user.organization_id
    from sqlalchemy import or_
    stmt = select(DocumentFileModel).where(
        DocumentFileModel.organization_id == org_id,
        or_(DocumentFileModel.status == "queued", DocumentFileModel.status == "error"),
        DocumentFileModel.chunks_count.in_([0, None]),
    )
    files = (await db.execute(stmt)).scalars().all()
    count = len(files)
    for f in files:
        await db.delete(f)
    await db.flush()
    logger.info("documents.queued_cleared", org=org_id, deleted=count)
    return {"deleted": count, "message": f"{count} stuck queued files removed. Re-upload to reprocess."}


@router.get("/files/{file_id}", response_model=DocumentFileOut)
async def get_file(
    file_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    org_id = user.organization_id
    stmt = select(DocumentFileModel).where(
        DocumentFileModel.id == file_id,
        DocumentFileModel.organization_id == org_id,
    )
    file = (await db.execute(stmt)).scalar_one_or_none()
    if not file:
        raise HTTPException(status_code=404, detail="Document not found")
    return file


@router.post("/files/{file_id}/process", status_code=status.HTTP_202_ACCEPTED)
async def process_single_file(
    file_id: str,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Re-trigger background processing for a single document (pending/failed/error)."""
    org_id = user.organization_id
    stmt = select(DocumentFileModel).where(
        DocumentFileModel.id == file_id,
        DocumentFileModel.organization_id == org_id,
    )
    doc = (await db.execute(stmt)).scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    if doc.status not in ("pending", "failed", "error"):
        raise HTTPException(status_code=409, detail=f"Document status is '{doc.status}', only pending/failed/error can be reprocessed")

    file_path = (doc.file_url or "").replace("upload://", "") if doc.file_url and doc.file_url.startswith("upload://") else doc.file_url
    content_type = "html" if (file_path or "").endswith(".html") else "pdf"

    # Reset to pending so the UI shows progress
    doc.status = "pending"
    doc.updated_at = datetime.now(UTC)

    background_tasks.add_task(
        process_document_background,
        doc_file_id=doc.id,
        file_path=file_path,
        content_type=content_type,
        org_id=org_id,
        supplier_id=doc.supplier_id,
        doc_type=doc.doc_type,
        company_name=doc.company_name,
        report_year=doc.report_year,
    )
    return {"queued": 1, "doc_file_id": doc.id}


@router.delete("/files/{file_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_file(
    file_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    org_id = user.organization_id
    stmt = select(DocumentFileModel).where(
        DocumentFileModel.id == file_id,
        DocumentFileModel.organization_id == org_id,
    )
    file = (await db.execute(stmt)).scalar_one_or_none()
    if not file:
        raise HTTPException(status_code=404, detail="Document not found")
    await db.delete(file)


# ── Bulk Ingest ───────────────────────────────────────────────────────────────

# Global cancellation flag per org — set by cancel-processing, cleared on start
_cancel_flags: dict[str, asyncio.Event] = {}


@router.post("/ingest-all")
async def ingest_all(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    org_id = user.organization_id
    stats = await ingest_all_active_sources(org_id, db)
    return {"organization_id": org_id, "stats": stats}


@router.post("/process-pending", status_code=status.HTTP_202_ACCEPTED)
async def process_pending(
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Queue all pending/failed uploaded documents for background processing."""
    org_id = user.organization_id

    # Clear any previous cancellation for this org
    flag = _cancel_flags.get(org_id)
    if flag:
        flag.clear()
    else:
        _cancel_flags[org_id] = asyncio.Event()

    stmt = select(DocumentFileModel).where(
        DocumentFileModel.organization_id == org_id,
        DocumentFileModel.status.in_(["pending", "failed", "error"]),
        DocumentFileModel.file_url.isnot(None),
    )
    docs = (await db.execute(stmt)).scalars().all()

    queued = 0
    for doc in docs:
        file_path = (doc.file_url or "").replace("upload://", "") if doc.file_url and doc.file_url.startswith("upload://") else doc.file_url
        content_type = "html" if (file_path or "").endswith(".html") else "pdf"
        background_tasks.add_task(
            _process_with_cancel,
            doc_file_id=doc.id,
            file_path=file_path,
            content_type=content_type,
            org_id=org_id,
            supplier_id=doc.supplier_id,
            doc_type=doc.doc_type,
            company_name=doc.company_name,
            report_year=doc.report_year,
            cancel_flag=_cancel_flags[org_id],
        )
        queued += 1

    return {"queued": queued, "organization_id": org_id}


async def _process_with_cancel(cancel_flag: asyncio.Event, **kwargs) -> None:
    """Wrapper: skip processing if cancel flag is set."""
    if cancel_flag.is_set():
        return
    await process_document_background(**kwargs)


@router.post("/cancel-processing", status_code=status.HTTP_200_OK)
async def cancel_processing(
    user: User = Depends(get_current_user),
):
    """Signal all pending background processing tasks to stop after the current document."""
    org_id = user.organization_id
    flag = _cancel_flags.get(org_id)
    if flag:
        flag.set()
    else:
        _cancel_flags[org_id] = asyncio.Event()
        _cancel_flags[org_id].set()
    return {"cancelled": True, "organization_id": org_id}


# ── Human-in-the-Loop Review ──────────────────────────────────────────────────

@router.get("/files/{file_id}/serve")
async def serve_file(
    file_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Serve the raw PDF/HTML file for in-browser viewing."""
    org_id = user.organization_id
    stmt = select(DocumentFileModel).where(
        DocumentFileModel.id == file_id,
        DocumentFileModel.organization_id == org_id,
    )
    doc = (await db.execute(stmt)).scalar_one_or_none()
    if not doc or not doc.file_url:
        raise HTTPException(status_code=404, detail="File not found")

    if doc.file_url.startswith("upload://"):
        # Legacy format: search in org upload dir
        filename = doc.file_url.replace("upload://", "")
        from shared.config import settings as _settings
        upload_dir = os.path.join(_settings.upload_storage_path, org_id)
        # Search for matching file (may have hash prefix)
        for candidate in os.listdir(upload_dir) if os.path.isdir(upload_dir) else []:
            if candidate.endswith(filename):
                file_path = os.path.join(upload_dir, candidate)
                break
        else:
            raise HTTPException(status_code=404, detail="Physical file not found")
    else:
        file_path = doc.file_url

    if not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="Physical file not found")

    media_type = "text/html" if file_path.endswith(".html") else "application/pdf"
    return FileResponse(file_path, media_type=media_type, filename=os.path.basename(file_path))


@router.get("/files/{file_id}/status-stream")
async def status_stream(
    file_id: str,
    token: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    """SSE stream that emits the document status every second until done/failed.

    EventSource cannot send custom headers, so the JWT is passed as `?token=<jwt>`.
    """
    import json
    from infrastructure.persistence.models.user import UserModel

    if not token:
        return StreamingResponse(
            iter([f"data: {json.dumps({'error': 'unauthorized'})}\n\n"]),
            media_type="text/event-stream",
        )
    try:
        payload = decode_token(token)
        user_id = payload.get("sub") or ""
        user_row = (await db.execute(
            select(UserModel.organization_id).where(UserModel.id == user_id)
        )).first()
        org_id = user_row[0] if user_row else ""
        if not org_id:
            raise ValueError("no org")
    except Exception:
        return StreamingResponse(
            iter([f"data: {json.dumps({'error': 'unauthorized'})}\n\n"]),
            media_type="text/event-stream",
        )

    async def generate():
        terminal = {"done", "failed", "error", "completed"}
        consecutive_errors = 0
        while True:
            try:
                from sqlalchemy import select as _select
                stmt = _select(
                    DocumentFileModel.status,
                    DocumentFileModel.chunks_count,
                    DocumentFileModel.error_msg,
                    DocumentFileModel.updated_at,
                ).where(
                    DocumentFileModel.id == file_id,
                    DocumentFileModel.organization_id == org_id,
                )
                row = (await db.execute(stmt)).first()
                if row is None:
                    yield f"data: {json.dumps({'error': 'not_found'})}\n\n"
                    return
                payload = {
                    "status": row[0],
                    "chunks_count": row[1] or 0,
                    "error_msg": row[2],
                    "updated_at": row[3].isoformat() if row[3] else None,
                }
                yield f"data: {json.dumps(payload)}\n\n"
                if row[0] in terminal:
                    return
                consecutive_errors = 0
            except Exception as exc:
                consecutive_errors += 1
                yield f"data: {json.dumps({'error': str(exc)})}\n\n"
                if consecutive_errors >= 3:
                    return
            await asyncio.sleep(1)

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@router.get("/files/{file_id}/review", response_model=ReviewDataOut)
async def get_file_review(
    file_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Full review payload: file metadata + chunks + metrics + signals + audit log."""
    org_id = user.organization_id
    stmt = select(DocumentFileModel).where(
        DocumentFileModel.id == file_id,
        DocumentFileModel.organization_id == org_id,
    )
    doc = (await db.execute(stmt)).scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    # Chunks
    chunk_stmt = select(RagDocumentModel).where(
        RagDocumentModel.document_file_id == file_id,
        RagDocumentModel.organization_id == org_id,
    ).order_by(RagDocumentModel.created_at)
    chunks = (await db.execute(chunk_stmt)).scalars().all()

    # Metrics
    metric_stmt = select(CompanyMetricModel).where(
        CompanyMetricModel.source_doc_id == file_id,
        CompanyMetricModel.organization_id == org_id,
    )
    metrics = (await db.execute(metric_stmt)).scalars().all()

    # Signals
    signal_stmt = select(CompanySignalModel).where(
        CompanySignalModel.source_doc_id == file_id,
        CompanySignalModel.organization_id == org_id,
    )
    signals = (await db.execute(signal_stmt)).scalars().all()

    # Audit log
    log_rows = (await db.execute(
        text("SELECT id, user_id, action, field, old_value, new_value, created_at FROM document_review_log WHERE doc_file_id = :fid ORDER BY created_at DESC LIMIT 100"),
        {"fid": file_id},
    )).mappings().all()

    has_pdf = bool(doc.file_url) and (
        os.path.isfile(doc.file_url) if not (doc.file_url or "").startswith("upload://") else True
    )

    return ReviewDataOut(
        id=doc.id,
        doc_type=doc.doc_type,
        company_name=doc.company_name,
        report_year=doc.report_year,
        title=doc.title,
        language=doc.language,
        pages=doc.pages,
        chunks_count=doc.chunks_count,
        esg_score=doc.esg_score,
        summary=doc.summary,
        status=doc.status,
        review_status=doc.review_status,
        review_notes=doc.review_notes,
        parsed_text=doc.parsed_text,
        extracted_kpis=doc.extracted_kpis,
        extracted_risks=doc.extracted_risks,
        extracted_targets=doc.extracted_targets,
        extracted_commitments=doc.extracted_commitments,
        has_pdf=has_pdf,
        copilot_hidden=bool(doc.copilot_hidden),
        classification_confidence=doc.classification_confidence,
        classification_alternatives=doc.classification_alternatives,
        classification_evidence=doc.classification_evidence or [],
        created_at=doc.created_at,
        updated_at=doc.updated_at,
        chunks=[ReviewChunkOut(id=c.id, content=c.content, chunk_level=c.chunk_level, doc_class=c.doc_class, page_number=c.page_number, excluded_from_index=bool(c.excluded_from_index)) for c in chunks],
        metrics=[ReviewMetricOut(id=m.id, metric_type=m.metric_type, value=float(m.value), unit=m.unit, year=m.year, period=m.period, confidence=m.confidence, confidence_pct=m.confidence_pct, page_number=m.page_number, scope=m.scope) for m in metrics],
        signals=[ReviewSignalOut(id=s.id, signal_type=s.signal_type, dimension=s.dimension, direction=s.direction, severity=s.severity, description=s.description, year=s.year) for s in signals],
        audit_log=[ReviewAuditEntry(id=str(r["id"]), user_id=r["user_id"], action=r["action"], field=r["field"], old_value=r["old_value"], new_value=r["new_value"], created_at=r["created_at"]) for r in log_rows],
    )


@router.get("/suppliers/{supplier_id}/audit-log", status_code=status.HTTP_200_OK)
async def get_supplier_audit_log(
    supplier_id: str,
    limit: int = 200,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Return all document_review_log entries across all documents of a supplier."""
    org_id = user.organization_id
    rows = (await db.execute(text("""
        SELECT l.id, l.doc_file_id, l.user_id, l.action, l.field, l.old_value, l.new_value, l.created_at,
               f.title, f.doc_type, f.report_year
        FROM document_review_log l
        JOIN document_files f ON f.id = l.doc_file_id
        WHERE f.supplier_id   = :sid
          AND f.organization_id = :org
        ORDER BY l.created_at DESC
        LIMIT :lim
    """), {"sid": supplier_id, "org": org_id, "lim": limit})).fetchall()

    return [
        {
            "id": str(r["id"]),
            "doc_file_id": r["doc_file_id"],
            "doc_title": r["title"] or r["doc_type"],
            "doc_type": r["doc_type"],
            "report_year": r["report_year"],
            "user_id": r["user_id"],
            "action": r["action"],
            "field": r["field"],
            "old_value": r["old_value"],
            "new_value": r["new_value"],
            "created_at": r["created_at"],
        }
        for r in rows
    ]


@router.patch("/files/{file_id}/classification", status_code=status.HTTP_200_OK)
async def update_classification(
    file_id: str,
    payload: ClassificationUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Update document classification metadata and log the change."""
    org_id = user.organization_id
    stmt = select(DocumentFileModel).where(
        DocumentFileModel.id == file_id,
        DocumentFileModel.organization_id == org_id,
    )
    doc = (await db.execute(stmt)).scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    changes: list[tuple[str, str | None, str | None]] = []
    if payload.doc_type is not None and payload.doc_type != doc.doc_type:
        changes.append(("doc_type", doc.doc_type, payload.doc_type))
        doc.doc_type = payload.doc_type
    if payload.company_name is not None and payload.company_name != doc.company_name:
        changes.append(("company_name", doc.company_name, payload.company_name))
        doc.company_name = payload.company_name
    if payload.report_year is not None and payload.report_year != doc.report_year:
        changes.append(("report_year", str(doc.report_year), str(payload.report_year)))
        doc.report_year = payload.report_year

    doc.updated_at = datetime.now(UTC)
    for field, old_val, new_val in changes:
        await db.execute(
            text("INSERT INTO document_review_log (id, doc_file_id, organization_id, user_id, action, field, old_value, new_value) VALUES (:id, :fid, :org, :uid, :action, :field, :old, :new)"),
            {"id": str(uuid.uuid4()), "fid": file_id, "org": org_id, "uid": user.id, "action": "update_classification", "field": field, "old": old_val, "new": new_val},
        )
    return {"updated": len(changes), "fields": [c[0] for c in changes]}


@router.patch("/files/{file_id}/kpis", status_code=status.HTTP_200_OK)
async def update_kpis(
    file_id: str,
    payload: KpiUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Replace extracted_kpis with manually edited values."""
    org_id = user.organization_id
    stmt = select(DocumentFileModel).where(
        DocumentFileModel.id == file_id,
        DocumentFileModel.organization_id == org_id,
    )
    doc = (await db.execute(stmt)).scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    import json as _json
    old_val = _json.dumps(doc.extracted_kpis or {}, ensure_ascii=False)
    doc.extracted_kpis = payload.kpis
    doc.updated_at = datetime.now(UTC)
    await db.execute(
        text("INSERT INTO document_review_log (id, doc_file_id, organization_id, user_id, action, field, old_value, new_value) VALUES (:id, :fid, :org, :uid, :action, :field, :old, :new)"),
        {"id": str(uuid.uuid4()), "fid": file_id, "org": org_id, "uid": user.id, "action": "update_kpis", "field": "extracted_kpis", "old": old_val, "new": _json.dumps(payload.kpis, ensure_ascii=False)},
    )
    return {"updated": True}


@router.post("/files/{file_id}/approve", status_code=status.HTTP_200_OK)
async def approve_document(
    file_id: str,
    payload: ApproveRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Mark document as human-approved."""
    org_id = user.organization_id
    stmt = select(DocumentFileModel).where(
        DocumentFileModel.id == file_id,
        DocumentFileModel.organization_id == org_id,
    )
    doc = (await db.execute(stmt)).scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    old_status = doc.review_status
    doc.review_status = "approved"
    if payload.notes:
        doc.review_notes = payload.notes
    doc.updated_at = datetime.now(UTC)
    await db.execute(
        text("INSERT INTO document_review_log (id, doc_file_id, organization_id, user_id, action, field, old_value, new_value) VALUES (:id, :fid, :org, :uid, :action, :field, :old, :new)"),
        {"id": str(uuid.uuid4()), "fid": file_id, "org": org_id, "uid": user.id, "action": "approve", "field": "review_status", "old": old_status, "new": "approved"},
    )
    return {"review_status": "approved"}


@router.post("/files/{file_id}/unapprove", status_code=status.HTTP_200_OK)
async def unapprove_document(
    file_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Revoke approval and set document back to draft."""
    org_id = user.organization_id
    stmt = select(DocumentFileModel).where(
        DocumentFileModel.id == file_id,
        DocumentFileModel.organization_id == org_id,
    )
    doc = (await db.execute(stmt)).scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    old_status = doc.review_status
    doc.review_status = "draft"
    doc.updated_at = datetime.now(UTC)
    await db.execute(
        text("INSERT INTO document_review_log (id, doc_file_id, organization_id, user_id, action, field, old_value, new_value) VALUES (:id, :fid, :org, :uid, :action, :field, :old, :new)"),
        {"id": str(uuid.uuid4()), "fid": file_id, "org": org_id, "uid": user.id, "action": "unapprove", "field": "review_status", "old": old_status, "new": "draft"},
    )
    return {"review_status": "draft"}


@router.delete("/chunks/{chunk_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_chunk(
    chunk_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Remove a chunk from the vector index."""
    org_id = user.organization_id
    stmt = select(RagDocumentModel).where(
        RagDocumentModel.id == chunk_id,
        RagDocumentModel.organization_id == org_id,
    )
    chunk = (await db.execute(stmt)).scalar_one_or_none()
    if not chunk:
        raise HTTPException(status_code=404, detail="Chunk not found")

    doc_file_id = chunk.document_file_id
    await db.delete(chunk)
    if doc_file_id:
        await db.execute(
            text("INSERT INTO document_review_log (id, doc_file_id, organization_id, user_id, action, field, old_value, new_value) VALUES (:id, :fid, :org, :uid, :action, :field, :old, :new)"),
            {"id": str(uuid.uuid4()), "fid": doc_file_id, "org": org_id, "uid": user.id, "action": "delete_chunk", "field": "chunk_id", "old": chunk_id, "new": None},
        )


@router.patch("/chunks/{chunk_id}", status_code=status.HTTP_200_OK)
async def update_chunk(
    chunk_id: str,
    payload: ChunkContentUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Edit chunk content (re-embedding happens async)."""
    org_id = user.organization_id
    stmt = select(RagDocumentModel).where(
        RagDocumentModel.id == chunk_id,
        RagDocumentModel.organization_id == org_id,
    )
    chunk = (await db.execute(stmt)).scalar_one_or_none()
    if not chunk:
        raise HTTPException(status_code=404, detail="Chunk not found")

    old_content = chunk.content[:500]
    chunk.content = payload.content
    chunk.embedding = None  # invalidate — re-embed on next retrieval pass
    doc_file_id = chunk.document_file_id
    if doc_file_id:
        await db.execute(
            text("INSERT INTO document_review_log (id, doc_file_id, organization_id, user_id, action, field, old_value, new_value) VALUES (:id, :fid, :org, :uid, :action, :field, :old, :new)"),
            {"id": str(uuid.uuid4()), "fid": doc_file_id, "org": org_id, "uid": user.id, "action": "edit_chunk", "field": "content", "old": old_content, "new": payload.content[:500]},
        )
    return {"updated": True, "embedding_invalidated": True}


@router.post("/files/{file_id}/test-retrieval")
async def test_retrieval(
    file_id: str,
    payload: TestRetrievalRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Run a similarity search against this document's chunks only."""
    org_id = user.organization_id
    stmt = select(DocumentFileModel).where(
        DocumentFileModel.id == file_id,
        DocumentFileModel.organization_id == org_id,
    )
    doc = (await db.execute(stmt)).scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    from infrastructure.llm.deps import get_org_pipeline_settings
    pipe = await get_org_pipeline_settings(org_id, db)
    effective_top_k = payload.top_k if payload.top_k is not None else pipe["top_k"]
    effective_min_sim = payload.min_sim if payload.min_sim is not None else pipe["similarity_threshold"]

    import asyncio as _aio
    import concurrent.futures as _cf
    loop = _aio.get_event_loop()
    with _cf.ThreadPoolExecutor(max_workers=1) as pool:
        query_vec: list[float] = await loop.run_in_executor(pool, embed_query, payload.query)

    vec_str = "[" + ",".join(str(v) for v in query_vec) + "]"
    rows = (await db.execute(
        text(f"""
            SELECT id, content, chunk_level, doc_class,
                   1 - (embedding <=> CAST('{vec_str}' AS vector)) AS similarity
            FROM rag_documents
            WHERE document_file_id = :fid
              AND organization_id = :org
              AND embedding IS NOT NULL
              AND 1 - (embedding <=> CAST('{vec_str}' AS vector)) >= :min_sim
            ORDER BY embedding <=> CAST('{vec_str}' AS vector)
            LIMIT :top_k
        """),
        {"fid": file_id, "org": org_id, "min_sim": effective_min_sim, "top_k": effective_top_k},
    )).mappings().all()

    return {
        "query": payload.query,
        "results": [
            {
                "chunk_id": r["id"],
                "similarity": round(float(r["similarity"]), 4),
                "chunk_level": r["chunk_level"],
                "doc_class": r["doc_class"],
                "content_preview": r["content"][:300],
            }
            for r in rows
        ],
    }


class ChunkSplitRequest(BaseModel):
    split_at: int  # character index in content where split occurs


@router.post("/chunks/{chunk_id}/split", status_code=status.HTTP_200_OK)
async def split_chunk(
    chunk_id: str,
    payload: ChunkSplitRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Split a chunk at a given character index, creating two new chunks."""
    org_id = user.organization_id
    chunk = (await db.execute(
        select(RagDocumentModel).where(
            RagDocumentModel.id == chunk_id,
            RagDocumentModel.organization_id == org_id,
        )
    )).scalar_one_or_none()
    if not chunk:
        raise HTTPException(status_code=404, detail="Chunk not found")

    content = chunk.content
    split_at = max(1, min(payload.split_at, len(content) - 1))
    part_a = content[:split_at].strip()
    part_b = content[split_at:].strip()

    if not part_a or not part_b:
        raise HTTPException(status_code=422, detail="Split would produce an empty chunk")

    # Update original chunk with part A
    old_content = content[:200]
    chunk.content = part_a
    chunk.embedding = None

    # Create new chunk for part B
    new_chunk = RagDocumentModel(
        id=str(uuid.uuid4()),
        organization_id=org_id,
        document_file_id=chunk.document_file_id,
        supplier_id=chunk.supplier_id,
        doc_type=chunk.doc_type,
        doc_class=chunk.doc_class,
        company_name=chunk.company_name,
        report_year=chunk.report_year,
        source_id=chunk.source_id,
        chunk_level=chunk.chunk_level,
        content=part_b,
        embedding=None,
    )
    db.add(new_chunk)

    if chunk.document_file_id:
        await db.execute(
            text("INSERT INTO document_review_log (id, doc_file_id, organization_id, user_id, action, field, old_value, new_value) VALUES (:id, :fid, :org, :uid, :action, :field, :old, :new)"),
            {"id": str(uuid.uuid4()), "fid": chunk.document_file_id, "org": org_id, "uid": user.id,
             "action": "split_chunk", "field": "content", "old": old_content, "new": f"split at {split_at}"},
        )

    return {"split": True, "chunk_a_id": chunk_id, "chunk_b_id": new_chunk.id}


class ChunkMergeRequest(BaseModel):
    other_chunk_id: str
    separator: str = "\n\n"


@router.post("/chunks/{chunk_id}/merge", status_code=status.HTTP_200_OK)
async def merge_chunks(
    chunk_id: str,
    payload: ChunkMergeRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Merge this chunk with another chunk (other is deleted)."""
    org_id = user.organization_id

    chunk_a = (await db.execute(
        select(RagDocumentModel).where(
            RagDocumentModel.id == chunk_id,
            RagDocumentModel.organization_id == org_id,
        )
    )).scalar_one_or_none()
    chunk_b = (await db.execute(
        select(RagDocumentModel).where(
            RagDocumentModel.id == payload.other_chunk_id,
            RagDocumentModel.organization_id == org_id,
        )
    )).scalar_one_or_none()

    if not chunk_a or not chunk_b:
        raise HTTPException(status_code=404, detail="One or both chunks not found")
    if chunk_a.document_file_id != chunk_b.document_file_id:
        raise HTTPException(status_code=422, detail="Chunks must belong to the same document")

    old_content = chunk_a.content[:200]
    chunk_a.content = chunk_a.content.strip() + payload.separator + chunk_b.content.strip()
    chunk_a.embedding = None
    await db.delete(chunk_b)

    if chunk_a.document_file_id:
        await db.execute(
            text("INSERT INTO document_review_log (id, doc_file_id, organization_id, user_id, action, field, old_value, new_value) VALUES (:id, :fid, :org, :uid, :action, :field, :old, :new)"),
            {"id": str(uuid.uuid4()), "fid": chunk_a.document_file_id, "org": org_id, "uid": user.id,
             "action": "merge_chunks", "field": "content", "old": old_content, "new": f"merged with {payload.other_chunk_id[:8]}"},
        )

    return {"merged": True, "surviving_chunk_id": chunk_id, "deleted_chunk_id": payload.other_chunk_id}


@router.patch("/chunks/{chunk_id}/exclude", status_code=status.HTTP_200_OK)
async def toggle_chunk_exclusion(
    chunk_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Toggle excluded_from_index for a chunk. Excluded chunks are skipped in retrieval."""
    org_id = user.organization_id
    chunk = (await db.execute(
        select(RagDocumentModel).where(
            RagDocumentModel.id == chunk_id,
            RagDocumentModel.organization_id == org_id,
        )
    )).scalar_one_or_none()
    if not chunk:
        raise HTTPException(status_code=404, detail="Chunk not found")

    chunk.excluded_from_index = not bool(chunk.excluded_from_index)
    if chunk.document_file_id:
        await db.execute(
            text("INSERT INTO document_review_log (id, doc_file_id, organization_id, user_id, action, field, old_value, new_value) VALUES (:id, :fid, :org, :uid, :action, :field, :old, :new)"),
            {"id": str(uuid.uuid4()), "fid": chunk.document_file_id, "org": org_id, "uid": user.id,
             "action": "toggle_exclude", "field": "excluded_from_index",
             "old": str(not chunk.excluded_from_index), "new": str(chunk.excluded_from_index)},
        )
    return {"excluded": chunk.excluded_from_index, "chunk_id": chunk_id}


@router.patch("/files/{file_id}/copilot-visibility", status_code=status.HTTP_200_OK)
async def toggle_copilot_visibility(
    file_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Toggle copilot_hidden — hides entire document from retrieval without deleting."""
    org_id = user.organization_id
    doc = (await db.execute(
        select(DocumentFileModel).where(
            DocumentFileModel.id == file_id,
            DocumentFileModel.organization_id == org_id,
        )
    )).scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    doc.copilot_hidden = not bool(doc.copilot_hidden)
    doc.updated_at = datetime.now(UTC)
    await db.execute(
        text("INSERT INTO document_review_log (id, doc_file_id, organization_id, user_id, action, field, old_value, new_value) VALUES (:id, :fid, :org, :uid, :action, :field, :old, :new)"),
        {"id": str(uuid.uuid4()), "fid": file_id, "org": org_id, "uid": user.id,
         "action": "toggle_copilot_visibility", "field": "copilot_hidden",
         "old": str(not doc.copilot_hidden), "new": str(doc.copilot_hidden)},
    )
    return {"copilot_hidden": doc.copilot_hidden, "file_id": file_id}


# ── Metrics ───────────────────────────────────────────────────────────────────

class MetricUpdate(BaseModel):
    value: float | None = None
    unit: str | None = None
    year: int | None = None
    period: str | None = None
    confidence: str | None = None
    page_number: int | None = None
    scope: str | None = None

class MetricCreate(BaseModel):
    metric_type: str
    value: float
    unit: str
    year: int
    period: str = "FY"
    confidence: str = "exact"


@router.patch("/metrics/{metric_id}", status_code=status.HTTP_200_OK)
async def update_metric(
    metric_id: str,
    payload: MetricUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Update a specific metric field."""
    org_id = user.organization_id
    metric = (await db.execute(
        select(CompanyMetricModel).where(
            CompanyMetricModel.id == metric_id,
            CompanyMetricModel.organization_id == org_id,
        )
    )).scalar_one_or_none()
    if not metric:
        raise HTTPException(status_code=404, detail="Metric not found")

    if payload.value is not None:
        metric.value = payload.value
    if payload.unit is not None:
        metric.unit = payload.unit
    if payload.year is not None:
        metric.year = payload.year
    if payload.period is not None:
        metric.period = payload.period
    if payload.confidence is not None:
        metric.confidence = payload.confidence
    if payload.page_number is not None:
        metric.page_number = payload.page_number
    if payload.scope is not None:
        metric.scope = payload.scope

    return {"id": metric_id, "updated": True}


@router.delete("/metrics/{metric_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_metric(
    metric_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Delete a metric."""
    org_id = user.organization_id
    metric = (await db.execute(
        select(CompanyMetricModel).where(
            CompanyMetricModel.id == metric_id,
            CompanyMetricModel.organization_id == org_id,
        )
    )).scalar_one_or_none()
    if not metric:
        raise HTTPException(status_code=404, detail="Metric not found")
    await db.delete(metric)


@router.post("/files/{file_id}/metrics", status_code=status.HTTP_201_CREATED)
async def add_metric(
    file_id: str,
    payload: MetricCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Add a new metric to a document."""
    org_id = user.organization_id
    doc = (await db.execute(
        select(DocumentFileModel).where(
            DocumentFileModel.id == file_id,
            DocumentFileModel.organization_id == org_id,
        )
    )).scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    metric = CompanyMetricModel(
        id=str(uuid.uuid4()),
        organization_id=org_id,
        company_name=doc.company_name or "unknown",
        metric_type=payload.metric_type,
        value=payload.value,
        unit=payload.unit,
        year=payload.year,
        period=payload.period,
        confidence=payload.confidence,
        source_doc_id=file_id,
    )
    db.add(metric)
    return ReviewMetricOut(
        id=metric.id,
        metric_type=metric.metric_type,
        value=float(metric.value),
        unit=metric.unit,
        year=metric.year,
        period=metric.period,
        confidence=metric.confidence,
        page_number=metric.page_number,
        scope=metric.scope,
    )


@router.get("/files/{file_id}/parse-layout", status_code=status.HTTP_200_OK)
async def get_parse_layout(
    file_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Return per-page bounding box layout data extracted during Docling parsing."""
    org_id = user.organization_id
    doc = (await db.execute(
        select(DocumentFileModel).where(
            DocumentFileModel.id == file_id,
            DocumentFileModel.organization_id == org_id,
        )
    )).scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"file_id": file_id, "pages": doc.pages, "layout": doc.parse_layout or {}}


class SandboxRequest(BaseModel):
    query: str


@router.post("/files/{file_id}/sandbox", status_code=status.HTTP_200_OK)
async def copilot_sandbox(
    file_id: str,
    payload: SandboxRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Test copilot retrieval against a single document (sandbox mode).
    Searches all chunks of the document regardless of excluded_from_index or copilot_hidden.
    Returns matching chunks with similarity scores + an LLM-generated answer.
    """
    org_id = user.organization_id

    doc = (await db.execute(
        select(DocumentFileModel).where(
            DocumentFileModel.id == file_id,
            DocumentFileModel.organization_id == org_id,
        )
    )).scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    query_vec = embed_query(payload.query)

    rows = (await db.execute(text("""
        SELECT id, content, page_number, excluded_from_index,
               1 - (embedding <=> CAST(:qv AS vector)) AS similarity
        FROM rag_documents
        WHERE document_file_id = :fid
          AND organization_id  = :org
          AND embedding IS NOT NULL
        ORDER BY embedding <=> CAST(:qv AS vector)
        LIMIT 8
    """), {"qv": str(query_vec), "fid": file_id, "org": org_id})).fetchall()

    chunks = [
        {
            "chunk_id": str(r.id),
            "content": r.content,
            "page_number": r.page_number,
            "excluded_from_index": bool(r.excluded_from_index),
            "similarity": round(float(r.similarity), 4),
        }
        for r in rows
    ]

    # P2-C: cross-corpus similar chunks from OTHER documents
    corpus_rows = (await db.execute(text("""
        SELECT rd.id, rd.content, rd.page_number,
               1 - (rd.embedding <=> CAST(:qv AS vector)) AS similarity,
               df.id AS doc_id, df.company_name, df.title, df.report_year, df.doc_type
        FROM rag_documents rd
        JOIN document_files df ON df.id = rd.document_file_id
        WHERE rd.organization_id = :org
          AND rd.document_file_id != :fid
          AND rd.embedding IS NOT NULL
          AND rd.excluded_from_index IS NOT TRUE
          AND (df.copilot_hidden IS NULL OR df.copilot_hidden IS FALSE)
          AND 1 - (rd.embedding <=> CAST(:qv AS vector)) >= 0.75
        ORDER BY rd.embedding <=> CAST(:qv AS vector)
        LIMIT 5
    """), {"qv": str(query_vec), "fid": file_id, "org": org_id})).fetchall()

    corpus_similar = [
        {
            "chunk_id": str(r.id),
            "content": r.content[:300],
            "page_number": r.page_number,
            "similarity": round(float(r.similarity), 4),
            "doc_id": str(r.doc_id),
            "company_name": r.company_name,
            "title": r.title,
            "report_year": r.report_year,
            "doc_type": r.doc_type,
        }
        for r in corpus_rows
    ]

    answer = None
    if chunks:
        context = "\n\n---\n\n".join(
            f"[Chunk {i+1}{' | S.' + str(c['page_number']) if c['page_number'] else ''}]\n{c['content'][:800]}"
            for i, c in enumerate(chunks)
        )
        llm = get_llm_provider()
        resp = await llm.complete(
            messages=[Message(role="user", content=payload.query)],
            system=(
                "You are a document assistant. Answer the question strictly based on the document excerpts below. "
                "Respond in the same language as the question. "
                "Cite the chunk number (e.g. [Chunk 2]) when referencing specific content. "
                "If the relevant information appears in any excerpt — even partially — quote it and explain. "
                "Only say the information is missing if it is genuinely absent from ALL excerpts.\n\n"
                f"DOCUMENT EXCERPTS:\n{context}"
            ),
            max_tokens=600,
        )
        answer = resp.content

    return {"query": payload.query, "answer": answer, "chunks": chunks, "corpus_similar": corpus_similar}


# ── Helper ────────────────────────────────────────────────────────────────────

async def _get_source_or_404(source_id: str, org_id: str, db: AsyncSession) -> DocumentSourceModel:
    stmt = select(DocumentSourceModel).where(
        DocumentSourceModel.id == source_id,
        DocumentSourceModel.organization_id == org_id,
    )
    source = (await db.execute(stmt)).scalar_one_or_none()
    if not source:
        raise HTTPException(status_code=404, detail="Source not found")
    return source


# ── Chunk Comments (P2-D) ─────────────────────────────────────────────────────

class ChunkCommentCreate(BaseModel):
    comment: str

@router.get("/chunks/{chunk_id}/comments", status_code=status.HTTP_200_OK)
async def list_chunk_comments(
    chunk_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    rows = (await db.execute(
        text("SELECT id, user_id, comment, created_at FROM chunk_comments WHERE chunk_id = :cid AND organization_id = :org ORDER BY created_at"),
        {"cid": chunk_id, "org": user.organization_id},
    )).mappings().all()
    return [{"id": str(r["id"]), "user_id": r["user_id"], "comment": r["comment"], "created_at": r["created_at"]} for r in rows]


@router.post("/chunks/{chunk_id}/comments", status_code=status.HTTP_201_CREATED)
async def create_chunk_comment(
    chunk_id: str,
    payload: ChunkCommentCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    comment_id = str(uuid.uuid4())
    await db.execute(
        text("INSERT INTO chunk_comments (id, chunk_id, organization_id, user_id, comment) VALUES (:id, :cid, :org, :uid, :comment)"),
        {"id": comment_id, "cid": chunk_id, "org": user.organization_id, "uid": user.id, "comment": payload.comment.strip()},
    )
    await db.commit()
    return {"id": comment_id, "chunk_id": chunk_id, "comment": payload.comment.strip()}


@router.delete("/chunks/{chunk_id}/comments/{comment_id}", status_code=status.HTTP_200_OK)
async def delete_chunk_comment(
    chunk_id: str,
    comment_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    await db.execute(
        text("DELETE FROM chunk_comments WHERE id = :id AND chunk_id = :cid AND organization_id = :org"),
        {"id": comment_id, "cid": chunk_id, "org": user.organization_id},
    )
    await db.commit()
    return {"deleted": comment_id}


# ── Export ────────────────────────────────────────────────────────────────────

@router.get("/files/{file_id}/export")
async def export_file_data(
    file_id: str,
    format: str = "xlsx",
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Export document metrics + metadata as JSON, CSV or Excel."""
    import csv
    import io
    import json as _json

    org_id = user.organization_id

    doc = (await db.execute(
        select(DocumentFileModel).where(
            DocumentFileModel.id == file_id,
            DocumentFileModel.organization_id == org_id,
        )
    )).scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    metrics = (await db.execute(
        select(CompanyMetricModel).where(
            CompanyMetricModel.source_doc_id == file_id,
            CompanyMetricModel.organization_id == org_id,
        ).order_by(CompanyMetricModel.metric_type)
    )).scalars().all()

    signals = (await db.execute(
        select(CompanySignalModel).where(
            CompanySignalModel.source_doc_id == file_id,
            CompanySignalModel.organization_id == org_id,
        )
    )).scalars().all()

    safe_name = (doc.company_name or doc.doc_type or "export").replace(" ", "_")[:40]
    year_part = f"_{doc.report_year}" if doc.report_year else ""

    # ── JSON ─────────────────────────────────────────────────────────────────
    if format == "json":
        payload = {
            "document": {
                "id": doc.id,
                "doc_type": doc.doc_type,
                "company_name": doc.company_name,
                "report_year": doc.report_year,
                "title": doc.title,
                "language": doc.language,
                "pages": doc.pages,
                "status": doc.status,
                "review_status": doc.review_status,
                "classification_confidence": doc.classification_confidence,
                "exported_at": datetime.now(UTC).isoformat(),
            },
            "metrics": [
                {
                    "metric_type": m.metric_type,
                    "value": float(m.value),
                    "unit": m.unit,
                    "year": m.year,
                    "period": m.period,
                    "confidence": m.confidence,
                    "confidence_pct": m.confidence_pct,
                    "scope": m.scope,
                    "page_number": m.page_number,
                }
                for m in metrics
            ],
            "signals": [
                {
                    "signal_type": s.signal_type,
                    "dimension": s.dimension,
                    "direction": s.direction,
                    "severity": s.severity,
                    "description": s.description,
                    "year": s.year,
                }
                for s in signals
            ],
        }
        content = _json.dumps(payload, ensure_ascii=False, indent=2)
        filename = f"{safe_name}{year_part}_export.json"
        return StreamingResponse(
            iter([content]),
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # ── CSV ──────────────────────────────────────────────────────────────────
    if format == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["metric_type", "value", "unit", "year", "period", "confidence", "confidence_pct", "scope", "page_number"])
        for m in metrics:
            writer.writerow([m.metric_type, float(m.value), m.unit, m.year, m.period, m.confidence, m.confidence_pct, m.scope, m.page_number])
        filename = f"{safe_name}{year_part}_metriken.csv"
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # ── Excel (default) ───────────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Sheet 1: Dokument-Übersicht
    ws_doc = wb.active
    ws_doc.title = "Dokument"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1E3A5F")
    doc_rows = [
        ("Feld", "Wert"),
        ("Dokument-Typ", doc.doc_type or ""),
        ("Unternehmen", doc.company_name or ""),
        ("Berichtsjahr", str(doc.report_year) if doc.report_year else ""),
        ("Titel", doc.title or ""),
        ("Sprache", doc.language or ""),
        ("Seiten", str(doc.pages) if doc.pages else ""),
        ("Status", doc.status or ""),
        ("Review-Status", doc.review_status or ""),
        ("Klassifizierungs-Konfidenz", f"{round((doc.classification_confidence or 0) * 100)}%" if doc.classification_confidence else ""),
        ("Anzahl Metriken", str(len(metrics))),
        ("Anzahl Signale", str(len(signals))),
        ("Exportiert am", datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")),
    ]
    for i, (k, v) in enumerate(doc_rows, start=1):
        ws_doc.cell(row=i, column=1, value=k)
        ws_doc.cell(row=i, column=2, value=v)
        if i == 1:
            for col in [1, 2]:
                cell = ws_doc.cell(row=i, column=col)
                cell.font = header_font
                cell.fill = header_fill
    ws_doc.column_dimensions["A"].width = 30
    ws_doc.column_dimensions["B"].width = 40

    # Sheet 2: Metriken
    ws_m = wb.create_sheet("Metriken")
    metric_headers = ["KPI", "Wert", "Einheit", "Jahr", "Periode", "Konfidenz", "Konfidenz %", "Scope", "Seite"]
    for col, h in enumerate(metric_headers, start=1):
        cell = ws_m.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for row, m in enumerate(metrics, start=2):
        conf_pct = m.confidence_pct if m.confidence_pct is not None else (95 if m.confidence == "exact" else 82 if m.confidence == "calculated" else 68)
        ws_m.append([m.metric_type, float(m.value), m.unit, m.year, m.period, m.confidence, conf_pct, m.scope, m.page_number])
    for col, width in zip(range(1, 10), [30, 14, 12, 8, 12, 14, 12, 14, 8]):
        ws_m.column_dimensions[get_column_letter(col)].width = width

    # Sheet 3: Signale
    ws_s = wb.create_sheet("Signale")
    signal_headers = ["Typ", "Dimension", "Richtung", "Schweregrad", "Beschreibung", "Jahr"]
    for col, h in enumerate(signal_headers, start=1):
        cell = ws_s.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for s in signals:
        ws_s.append([s.signal_type, s.dimension, s.direction, s.severity, s.description, s.year])
    for col, width in zip(range(1, 7), [20, 18, 14, 14, 60, 8]):
        ws_s.column_dimensions[get_column_letter(col)].width = width
    ws_s.column_dimensions["E"].alignment = Alignment(wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{safe_name}{year_part}_export.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
